from openpyxl import load_workbook
wb = load_workbook("sample_5.xlsx")
ws = wb.active

# 영어 성적이 80점 이상인 학생 선별
for row in ws.iter_rows(min_row=2):
    # 번호, 영어, 수학
    if int(row[1].value) > 80: # 값을 비교하기 위해 int사용
        print(row[0].value, "번 학생은 영어 천재")

# 값 변경
for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == "영어":
            cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")