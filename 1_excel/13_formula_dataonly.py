from openpyxl import load_workbook
# # 수식 그대로 가져옴
# wb = load_workbook("sample_formula.xlsx")
# ws = wb.active
#
# for row in ws.values:
#     for cell in row:
#         print(cell)

# 실제 데이터를 가져옴 - 수식 부분을 none라고 뜸
wb = load_workbook("sample_formula.xlsx", data_only=True) # 데이터 여는 속성 true로
ws = wb.active
# evaluate 되지 않은 상태의 데이터는 None이라고 표시
# 현재 엑셀파일을 열면 저장하겠냐고 물음 -> 열기 전엔 수식만 있는데 열고나서 계산이 되었으로
# 그것을 저장하면 수식결과된 값이 저장되어서 값 잘나옴
# 수식이 그대로 들어간거라 한 번 저장해야 계산이 되서 계산 값이 나옴
for row in ws.values:
    for cell in row:
        print(cell)