from openpyxl import load_workbook
wb = load_workbook("sample_5.xlsx")
ws = wb.active

# 번호 영어 수학 ->
# 번호 (국어) 영어 수학

# # ws.move_range("B1:C11", rows=0, cols=1) #0: 그대로
# ws["B1"].value = "국어" #B1 셀이 '국어'입력

# 번호 영어 수학 -> 수학 위치를 영어 위치로 옮기기
ws.move_range("C1:C11", rows=5, cols=-1)

wb.save("sample_korean.xlsx")
