from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
# 폰트, 테두리, 배경, 정렬 설정하기 위해서
from openpyxl import load_workbook
wb = load_workbook("sample_5.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"] # 번호
b1 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

# 열 너비 지정 - A 열의 너비를 5로 설정
ws.column_dimensions["A"].width = 5

# 행 높이 지정 - 1행의 높이를 50으로 지정
ws.row_dimensions[1].height = 50

# 스타일 적용
# 폰트변경
a1.font = Font(color="FF0000", italic=True, bold=True) # 글자 빨갛게, 이텔릭, 두껍게 적용
b1.font = Font(color="CC33FF", name="Arial", strike=True) # 글자 분홍색, 폰트체, 취소선 적용
c1.font = Font(color="0000FF", size=20, underline="single") # 글자 파란색, 사이즈, 한줄밑줄 적용

# 테두리선 적용
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

# 셀 배경 적용 - 90점 넘는 셀 초록색
for row in ws.rows: #전체 row가져오기
    for cell in row:
        # 각 cell에 대해서 정렬
        cell.alignment = Alignment(horizontal="center", vertical="center")
        # center, left, right, top, bottom
        if cell.column == 1: #A번호열은 제외
            continue

        # cell이 정수형 데이터고 90점보다 높으면
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="00FF00", fill_type="solid") #배경을 초록색으로 설정
            cell.font = Font(color="FF0000") #폰트 색상 변경

# 틀 고정
ws.freeze_panes = "B2"   #B2 기준으로 틀 고정

wb.save("sample_style.xlsx")