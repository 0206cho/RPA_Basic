from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기 -  리스트나 튜플 형식으로
ws.append(["번호", "영어", "수학"]) #A, B, C
for i in range(1, 11): #10개 데이터 넣기
    ws.append([i, randint(0,100), randint(0,100)])

# B열의 영어column만 가지고 오기
# col_B = ws["B"]
# print(col_B)

# 특정 열의 셀 데이터 가져오기
# for cell in col_B:
#     print(cell.value)

# 영어와 수학 컬럼 같이 가져오기
# col_range = ws["B:C"] # B부터 C
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# ROW도 가능 - 첫번째 행 가져오기
# row_title = ws[1]
# for cell in row_title:
#     print(cell.value)

# 여러 줄의 행 가져오기
# row_range = ws[2:6]
# 1번째 줄인 title을 제외하고 2번째 줄에서 6번째 줄까지 가지고 오기 2,3,4,5,6가져옴.
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# 행의 번호를 잘 모를 때
# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print()

# from openpyxl.utils.cell import coordinate_from_string  # 셀 정보 가져오기
# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.coordinate, end=" ")
#         xy = coordinate_from_string(cell.coordinate)
#         #A250을 A / 250 으로 끊어줘서 행이나 열 번호가 클 때 도움
#         # print(xy, end=" ")
#         print(xy[0], end="")  # A
#         print(xy[1], end=" ") # 1
#         # 위 두줄과 같음 print(cell.coordinate, end=" ")
#     print()

# 전체 rows - 모든 줄
#print(ws.rows)로 하면 알수 없는 정보가 뜸
# print(tuple(ws.rows)) # 괄호 묶음 단위가 행, 밑 두줄이랑 같음
# for row in tuple(ws.rows):
#     print(row)

# 일부 row
# for row in tuple(ws.rows):
#     print(row[1].value) # 인덱스1에 있는 값 찍어옴

# 다른 방법
# print()
# for row in ws.iter_rows():  # 전체 row 반복해서 가져옴
#     print(row[1].value)

# 타이틀만 출력
# for column in tuple(ws.columns):
#     print(column[0].value)

# 다른 방법
# for column in ws.iter_cols(): #전체 row
#     print(column[0].value)

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
#     print(row[0].value, row[1].value)

# 데이터 출력되는 모양 확인 - 범위 지정안할 시 min젤 작은, max젤 큰 자동으로 됨.
# 위에서 아래로
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
#     print(row)
# 왼쪽에서 오른쪽으로
# for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
#     print(col)

wb.save("sample_5.xlsx")