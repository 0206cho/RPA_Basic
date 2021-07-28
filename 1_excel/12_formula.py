import datetime # 날짜정보사용하기 위해서
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws["A1"] = datetime.datetime.today() #오늘 날짜 정보
ws["A2"] = "=SUM(1,2,3)"  # 합계 1+2+3=6
ws["A3"] = "=AVERAGE(1,2,3)" # 평균 2

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)" #30

wb.save("sample_formula.xlsx")
