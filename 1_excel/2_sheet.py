from openpyxl import Workbook
wb = Workbook() # 새로운 워크북 생성
# wb.active = 현재 활성화 된 시트 가져오기
ws = wb.create_sheet() # 현재 활성화된 시트 1번 뒤에 기본이름으로 새로운 시트 생성
ws.title = "MySheet" # shett이름 변경
ws.sheet_properties.tabColor = "b797e7" # RGB 형태로 값을 넣어주면 시트 탭 색상 변경

# sheet, mysheet, yoursheet 순으로 생성됨.
ws1 = wb.create_sheet("YourSheet") # 생성하면서 타이틀 넣어주기 - 주어진 이름으로 sheet생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 인덱스에 시트 생성

# 시트명으로 접근 할 떄 ws1. 접근해도 되지만 다음처럼
new_ws = wb["NewSheet"] # 딕셔너리 형태로 시트에 접근 가능
# new_ws..

# 엑셀 문서안의 모든 시트를 확인하기 위해서는 워크북.sheetnames
print(wb.sheetnames) # 모든 시트 이름 확인

# 시트 안에 있는 내용 복사
new_ws["A1"] = "Test" # A1셀의 값을 Test로 입력
target = wb.copy_worksheet(new_ws) # new_ws를 복사해서 target에 저장
target.title = "Copied Sheet" # target의 워크 시트 이름을 copied Sheet로 변경

wb.save("samples.xlsx") # 파일 저장