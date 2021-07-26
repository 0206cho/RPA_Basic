from openpyxl import Workbook

wb = Workbook()
ws = wb.active  # 현재 활성화된 Sheet가져옴
ws.title = "NadoSheet"

# A1 셀이 1 이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

# 셀 값 읽어오기
print(ws["A1"])  # A1 셀의 정보를 출력 -> 값이 아닌 셀 객체 정보를 출력
print(ws["A1"].value)  # A1셀의 값을 출력
print(ws["A10"].value)  # A10셀의 값을 출력 -> 값이 없을 때는 'None'를 출력

# column = A, B, C, ... -> A=1, B=2, ...
# row = 1, 2, 3,...
print(ws.cell(column=1, row=1).value)  # = ws["A1"].value
print(ws.cell(column=2, row=1).value)  # = ws["B1"].value

# 값 바로 입력
c = ws.cell(column=3, row=1, value=10)  # ws["C1"] = 10 을 c에 저장
print(c.value)  # ws["C1"].value

# 반복문 사용해 랜덤 데이터 넣기
from random import *  # 원래 상단에 넣어야 하는데 이해를 돕기 위해 여기에 삽입

'''
for x in range(1, 11): # 10개 row-행
    for y in range (1, 11): # 10개 column-열
        ws.cell(row=x, column=y, value=randint(0,100)) # 0~100 사이의 숫자
'''
index = 1
for x in range(1, 11):  # 10개 row-행
    for y in range(1, 11):  # 10개 column-열
        ws.cell(row=x, column=y, value=index)  # 좌표값처럼 순서대로 입력됨
        index += 1
        # 1 2 3 4 5
        # 6 7 8 9 10

wb.save("sample.xlsx")